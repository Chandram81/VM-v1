import openpyxl
import time
import os

def load_sheet_names(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        return workbook.sheetnames
    except Exception as e:
        raise RuntimeError(f"An error occurred while loading sheet names: {e}")

def load_columns(file_path, sheet_name):
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook[sheet_name]
        column_names = [cell.value for cell in sheet[1]]
        return column_names
    except Exception as e:
        raise RuntimeError(f"An error occurred while loading columns for sheet {sheet_name}: {e}")

import openpyxl
import time
import os

def combine_sheets(file_path, selected_sheets):
    try:
        workbook = openpyxl.load_workbook(file_path)
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Combined"
        
        # Dictionary to hold all unique column names
        column_names = set()

        # Collect all selected columns across sheets
        column_mapping = {}
        for sheet_name, selected_columns in selected_sheets:
            if not selected_columns:
                continue

            sheet = workbook[sheet_name]
            columns = [cell.value for cell in sheet[1] if cell.value is not None]
            column_indices = {col: idx + 1 for idx, col in enumerate(columns)}
            
            for _, col in selected_columns:
                if col not in column_mapping:
                    column_mapping[col] = len(column_names)
                    column_names.add(col)

        # Convert set to sorted list
        all_selected_columns = sorted(column_names)

        # Write headers
        new_sheet.append(all_selected_columns)

        # Write data rows
        for sheet_name, selected_columns in selected_sheets:
            if not selected_columns:
                continue

            sheet = workbook[sheet_name]
            columns = [cell.value for cell in sheet[1] if cell.value is not None]
            column_indices = {col: idx + 1 for idx, col in enumerate(columns)}
            
            # Get column indices for selected columns
            selected_column_indices = [column_indices[col] for _, col in selected_columns if col in column_indices]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                new_row = [''] * len(all_selected_columns)
                for idx in selected_column_indices:
                    header_name = columns[idx - 1]
                    if header_name in column_mapping:
                        new_row[column_mapping[header_name]] = row[idx - 1]
                new_sheet.append(new_row)

        # Adjust column widths to fit content
        for col in new_sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            new_sheet.column_dimensions[column].width = adjusted_width

        # Generate a Unix timestamp
        timestamp = int(time.time())

        # Convert the file path to a new Excel file path with timestamp
        base_name = os.path.basename(file_path)
        new_excel_file_path = os.path.join(
            os.path.dirname(file_path),
            base_name.replace('.xlsx', f'_combined_{timestamp}.xlsx')
        )

        # Save the new workbook
        new_workbook.save(new_excel_file_path)
        return new_excel_file_path
    except Exception as e:
        raise RuntimeError(f"An error occurred while combining the sheets: {e}")


